import os
import time
import smtplib
import requests
from datetime import date
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
APIFY_API_KEY   = os.environ["APIFY_API_KEY"]
EMAIL_SENDER    = os.environ["EMAIL_SENDER"]
EMAIL_PASSWORD  = os.environ["EMAIL_PASSWORD"]   # Gmail App Password (16 chars, no spaces)
EMAIL_RECIPIENT = os.environ["EMAIL_RECIPIENT"]

# ─── VERIFIED ACTOR IDS ───────────────────────────────────────────────────────
ACTOR_LINKEDIN    = "curious_coder/linkedin-jobs-scraper"
ACTOR_INDEED      = "borderline/indeed-scraper"
ACTOR_NAUKRI      = "muhammetakkurtt/naukri-job-scraper"
ACTOR_INTERNSHALA = "bareezh_codes/internshala-scrapper"

# ─── LINKEDIN SEARCH URLS ─────────────────────────────────────────────────────
# Built from linkedin.com/jobs/search — f_E=1 = Entry level, f_WT=2 = Remote
LINKEDIN_URLS = [
    "https://www.linkedin.com/jobs/search/?keywords=software%20engineer%20fresher&location=India&f_E=1",
    "https://www.linkedin.com/jobs/search/?keywords=junior%20software%20developer&location=India&f_E=1",
    "https://www.linkedin.com/jobs/search/?keywords=software%20engineer&location=Worldwide&f_WT=2&f_E=1",
]


def run_actor(actor_id: str, run_input: dict, timeout_secs: int = 300) -> list:
    """Start an Apify actor run, poll until done, return dataset items."""
    url = f"https://api.apify.com/v2/acts/{actor_id}/runs?token={APIFY_API_KEY}"
    resp = requests.post(url, json=run_input, timeout=30)

    if resp.status_code == 404:
        raise RuntimeError(f"Actor '{actor_id}' not found (404). Check actor name on apify.com.")
    resp.raise_for_status()

    run_id = resp.json()["data"]["id"]
    status_url = f"https://api.apify.com/v2/actor-runs/{run_id}?token={APIFY_API_KEY}"

    deadline = time.time() + timeout_secs
    while time.time() < deadline:
        time.sleep(10)
        status_resp = requests.get(status_url, timeout=15).json()
        status = status_resp["data"]["status"]
        if status in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
            break

    if status != "SUCCEEDED":
        print(f"  ⚠ Actor {actor_id} ended with status: {status}")
        return []

    dataset_id = status_resp["data"]["defaultDatasetId"]
    items_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items?token={APIFY_API_KEY}&format=json&limit=300"
    return requests.get(items_url, timeout=30).json()


def normalise(raw: dict, source: str) -> dict | None:
    def pick(*keys):
        for k in keys:
            v = raw.get(k)
            if v and str(v).strip() not in ("", "N/A", "null", "None", "nan"):
                return str(v).strip()
        return "N/A"

    company = pick("company", "companyName", "employer", "company_name", "organizationName", "hiringOrganization")
    role    = pick("title", "jobTitle", "position", "role", "positionTitle", "name")
    ctc     = pick("salary", "ctc", "salaryRange", "compensation", "stipend", "pay", "salary_range", "salaryText")
    link    = pick("url", "applyUrl", "jobUrl", "link", "applyLink", "jobLink", "externalApplyLink", "jobPostingUrl")
    loc     = pick("location", "jobLocation", "city", "place", "jobCity")
    posted  = pick("postedAt", "datePosted", "publishedAt", "date", "postedDate")

    if role == "N/A" or link == "N/A":
        return None

    return {
        "Source":       source,
        "Company":      company,
        "Role":         role,
        "CTC / Salary": ctc,
        "Apply Link":   link,
        "Location":     loc,
        "Posted":       posted,
    }


# ─── SCRAPERS ─────────────────────────────────────────────────────────────────

def scrape_linkedin() -> list:
    """LinkedIn: pass pre-built search URLs directly (correct schema)."""
    results = []
    print(f"  LinkedIn: scraping {len(LINKEDIN_URLS)} search URLs")
    try:
        raw = run_actor(ACTOR_LINKEDIN, {
            "urls": LINKEDIN_URLS,
            "count": 50,
            "scrapeCompany": False,
        })
        for item in raw:
            n = normalise(item, "LinkedIn")
            if n:
                results.append(n)
        print(f"  LinkedIn: {len(results)} jobs found")
    except Exception as e:
        print(f"  ⚠ LinkedIn failed: {e}")
    return results


def scrape_indeed() -> list:
    """Indeed: borderline/indeed-scraper."""
    results = []
    queries = [
        ("software engineer fresher", "India"),
        ("junior developer entry level", "India"),
        ("software developer fresher", "Remote"),
    ]
    for keyword, location in queries:
        print(f"  Indeed: '{keyword}' | {location}")
        try:
            raw = run_actor(ACTOR_INDEED, {
                "keyword": keyword,
                "location": location,
                "maxItems": 25,
            })
            for item in raw:
                n = normalise(item, "Indeed")
                if n:
                    results.append(n)
        except Exception as e:
            print(f"  ⚠ Indeed '{keyword}' failed: {e}")
    print(f"  Indeed: {len(results)} jobs found")
    return results


def scrape_naukri() -> list:
    """Naukri: muhammetakkurtt/naukri-job-scraper."""
    results = []
    queries = ["software engineer fresher", "junior developer", "entry level IT"]
    for keyword in queries:
        print(f"  Naukri: '{keyword}'")
        try:
            raw = run_actor(ACTOR_NAUKRI, {
                "keyword": keyword,
                "experience": "0",
                "maxJobs": 25,
            })
            for item in raw:
                n = normalise(item, "Naukri")
                if n:
                    results.append(n)
        except Exception as e:
            print(f"  ⚠ Naukri '{keyword}' failed: {e}")
    print(f"  Naukri: {len(results)} jobs found")
    return results


def scrape_internshala() -> list:
    """Internshala: bareezh_codes/internshala-scrapper."""
    results = []
    print("  Internshala: scraping fresher jobs")
    try:
        raw = run_actor(ACTOR_INTERNSHALA, {
            "category": "software-development",
            "workFromHome": False,
            "maxItems": 50,
        })
        for item in raw:
            n = normalise(item, "Internshala")
            if n:
                results.append(n)
        print(f"  Internshala: {len(results)} jobs found")
    except Exception as e:
        print(f"  ⚠ Internshala failed: {e}")
    return results


def deduplicate(jobs: list) -> list:
    seen, out = set(), []
    for j in jobs:
        key = (j["Company"].lower(), j["Role"].lower())
        if key not in seen:
            seen.add(key)
            out.append(j)
    return out


# ─── EXCEL BUILDER ────────────────────────────────────────────────────────────
SOURCE_COLORS = {
    "LinkedIn":    "0A66C2",
    "Indeed":      "2164F3",
    "Naukri":      "FF7555",
    "Internshala": "006BFF",
}

def build_excel(jobs: list, filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fresher Jobs"

    headers = ["#", "Source", "Company", "Role", "CTC / Salary", "Location", "Posted", "Apply Link"]
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30

    for i, job in enumerate(jobs, 1):
        row = i + 1
        alt_fill = PatternFill("solid", fgColor="F8FAFC") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        data = [i, job["Source"], job["Company"], job["Role"],
                job["CTC / Salary"], job["Location"], job["Posted"], job["Apply Link"]]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = alt_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in (4, 8)))
            if col == 2:
                color = SOURCE_COLORS.get(str(val), "6B7280")
                cell.font = Font(name="Arial", size=10, bold=True, color=color)
            elif col == 8:
                cell.font = Font(name="Arial", size=10, color="2563EB", underline="single")
                if str(val).startswith("http"):
                    cell.hyperlink = str(val)
            else:
                cell.font = Font(name="Arial", size=10)

    widths = [5, 13, 22, 35, 20, 18, 14, 45]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Job Scrape Summary"
    ws2["A1"].font = Font(bold=True, size=14, name="Arial")
    ws2["A3"], ws2["B3"] = "Date", str(date.today())
    ws2["A4"], ws2["B4"] = "Total Jobs", len(jobs)
    ws2["A5"], ws2["B5"] = "Sources", "LinkedIn, Indeed, Naukri, Internshala"
    ws2["A7"] = "Breakdown by Source"
    ws2["A7"].font = Font(bold=True, name="Arial")
    source_counts = {}
    for j in jobs:
        source_counts[j["Source"]] = source_counts.get(j["Source"], 0) + 1
    for r, (src, cnt) in enumerate(source_counts.items(), 8):
        ws2.cell(row=r, column=1, value=src)
        ws2.cell(row=r, column=2, value=cnt)
    for col in ["A", "B"]:
        ws2.column_dimensions[col].width = 25

    wb.save(filepath)
    print(f"✅ Saved: {filepath}")


# ─── EMAIL ────────────────────────────────────────────────────────────────────
def send_email(filepath: str, job_count: int):
    today = date.today().strftime("%d %b %Y")
    msg = MIMEMultipart()
    msg["From"]    = EMAIL_SENDER
    msg["To"]      = EMAIL_RECIPIENT
    msg["Subject"] = f"🧑‍💻 Fresher Software Jobs — {today} ({job_count} jobs)"

    body = f"""Hi,

Here is your daily fresher software job report for {today}.

📊 Total jobs found: {job_count}
🔍 Sources: LinkedIn, Indeed, Naukri, Internshala
📍 Locations: India + Remote

Open the attached Excel file to view all listings with direct apply links.

Good luck! 🚀
"""
    msg.attach(MIMEText(body, "plain"))

    with open(filepath, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(filepath)}")
    msg.attach(part)

    print("📧 Connecting to Gmail SMTP...")
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, EMAIL_RECIPIENT, msg.as_string())
        print(f"✅ Email sent to {EMAIL_RECIPIENT}")
    except smtplib.SMTPAuthenticationError:
        print("\n❌ Gmail authentication failed!")
        print("   Your EMAIL_PASSWORD secret must be a Gmail APP PASSWORD, not your regular password.")
        print("   Steps to fix:")
        print("   1. Go to myaccount.google.com/security")
        print("   2. Enable 2-Step Verification if not already done")
        print("   3. Search 'App Passwords' → Create one for Mail")
        print("   4. Copy the 16-character password (e.g. abcd efgh ijkl mnop)")
        print("   5. Update your GitHub secret EMAIL_PASSWORD with this value (remove spaces)")
        raise


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    today    = date.today().strftime("%Y-%m-%d")
    filepath = os.path.join(os.path.dirname(__file__), f"fresher_jobs_{today}.xlsx")

    print("🔍 Starting job scrape...\n")
    all_jobs = []

    print("[1/4] LinkedIn")
    all_jobs += scrape_linkedin()

    print("[2/4] Indeed")
    all_jobs += scrape_indeed()

    print("[3/4] Naukri")
    all_jobs += scrape_naukri()

    print("[4/4] Internshala")
    all_jobs += scrape_internshala()

    jobs = deduplicate(all_jobs)
    print(f"\n✅ Total unique jobs: {len(jobs)}\n")

    build_excel(jobs, filepath)
    send_email(filepath, len(jobs))


if __name__ == "__main__":
    main()
